from openpyxl import Workbook


class ExcelExportService:

    @staticmethod
    def generate(calculations):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Loan History"

        headers = [
            "Loan Amount",
            "Interest Rate (%)",
            "Tenure",
            "Monthly EMI",
            "Interest Paid",
            "Total Payment",
            "Created On",
        ]

        for column, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=column).value = header

        row = 2

        for item in calculations:
            sheet.cell(row=row, column=1).value = float(item.loan_amount)
            sheet.cell(row=row, column=2).value = float(item.annual_interest_rate)
            sheet.cell(row=row, column=3).value = f"{item.loan_tenure} {item.tenure_type}"
            sheet.cell(row=row, column=4).value = float(item.monthly_emi)
            sheet.cell(row=row, column=5).value = float(item.total_interest)
            sheet.cell(row=row, column=6).value = float(item.total_payment)
            sheet.cell(row=row, column=7).value = item.created_at.strftime("%d-%m-%Y %H:%M")

            row += 1

        return workbook